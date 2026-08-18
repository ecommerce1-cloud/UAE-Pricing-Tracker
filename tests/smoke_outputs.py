"""Offline check of the input parsing and output writers.

Exercises everything except the network: spreadsheet round-trip, Excel workbook
generation, and the standalone HTML dashboard (including that logos really are
inlined, since a file:// page cannot fetch them).
"""

import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scraper.inputs import load_products, write_template  # noqa: E402
from scraper.outputs import long_rows, write_excel, write_html  # noqa: E402
from scraper.result import classify_failure, ok, unavailable  # noqa: E402
from scraper.zones import ZONES  # noqa: E402

failures = []


def check(label, condition):
    print(("  PASS  " if condition else "  FAIL  ") + label)
    if not condition:
        failures.append(label)


def main():
    tmp = Path(tempfile.mkdtemp())

    print("spreadsheet round-trip")
    template = tmp / "products.xlsx"
    write_template(template)
    check("template file created", template.exists())
    products = load_products(template)
    check("template row parsed", len(products) == 1)
    barcode = next(iter(products))
    check("barcode kept as text, no float artefact", barcode == "5056141881928")
    refs = products[barcode]["platforms"]
    check("bare ASIN mapped to both Amazon tiers", "amazon_core" in refs and "amazon_fast" in refs)
    check("ASIN parsed as id not url", refs["amazon_core"].get("asin") == "B092BTLHZX")
    check("noon URL parsed as url", str(refs["noon_retail"].get("url", "")).startswith("http"))
    check("blank talabat column skipped", "talabat_mart" not in refs)

    print("failure classification")
    check("cloudflare detected", classify_failure("Just a moment...", "noon").startswith("blocked:"))
    check("amazon interstitial detected", classify_failure("Amazon.ae", "amazon").startswith("blocked:"))
    check("real page not misreported", not classify_failure("Hotpack Bags", "amazon").startswith("blocked:"))

    print("output writers")
    checked_at = datetime.now(timezone.utc).isoformat()
    results = {
        barcode: {
            "amazon_core": ok(23.65, "dom"),
            "amazon_fast": unavailable("not fast-delivery eligible"),
            "noon_retail": ok(22.45, "json_ld"),
            "noon_minutes": {
                "zones": {
                    z["id"]: (ok(31.05 + i * 0.5, "catalog_api") if i < 4 else unavailable("blocked: Cloudflare bot challenge"))
                    for i, z in enumerate(ZONES)
                }
            },
            "talabat_mart": {"zones": {z["id"]: unavailable("blocked: Cloudflare bot challenge") for z in ZONES}},
        }
    }

    rows = long_rows(products, results, checked_at)
    # 3 national platforms + 2 zoned platforms x 5 zones
    check("long rows cover every zone", len(rows) == 3 + 2 * len(ZONES))

    xlsx = tmp / "prices.xlsx"
    write_excel(xlsx, products, results, checked_at)
    check("excel written", xlsx.exists() and xlsx.stat().st_size > 4000)

    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    check("both sheets present", wb.sheetnames == ["Summary", "By Zone"])
    summary = wb["Summary"]
    check("summary has a row per barcode", summary.max_row == 2 + len(products))
    values = [c.value for c in summary[3]]
    check("barcode in leftmost column", values[0] == barcode)
    check("national price rendered", values[2] == "23.65")
    check("zone range rendered as min-max", isinstance(values[5], str) and "-" in values[5])

    html = tmp / "dashboard.html"
    write_html(html, products, results, checked_at)
    text = html.read_text(encoding="utf-8")
    check("html written", html.exists())
    check("logos inlined as data uris", text.count("data:image/svg+xml;base64,") == 5)
    check("no external references", "src=\"http" not in text and "href=\"http" not in text)
    check("unresolved template tokens gone", "__" not in text.replace("__DATE__PLACEHOLDER", ""))
    check("blocked cells marked with warning glyph", "&#9888;" in text)
    check("zone breakdown present", "zbreak" in text)
    check("export data embedded", "const ROWS" in text)

    print("")
    if failures:
        print(str(len(failures)) + " CHECK(S) FAILED:")
        for f in failures:
            print("  - " + f)
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
