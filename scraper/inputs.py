"""Read the barcode list from a spreadsheet.

Hand-editing JSON for 100 products across 4 platforms is unworkable, so the
input is a spreadsheet: one row per barcode, one column per platform holding
either the full product URL or the bare ID (ASIN/SKU).
"""

import csv
from pathlib import Path

# Accepted spellings for each input column, matched case/space-insensitively.
COLUMN_ALIASES = {
    "barcode": ("barcode", "ean", "upc", "barcode (ean/upc)"),
    "name": ("name", "product", "product name", "description"),
    "amazon": ("amazon", "amazon.ae", "amazon url", "amazon asin", "asin"),
    "noon": ("noon", "noon url", "noon sku", "noon retail"),
    "noon_minutes": ("noon minutes", "noon_minutes", "minutes", "noon minutes url"),
    "talabat": ("talabat", "talabat mart", "talabat url", "talabat_mart"),
}

TEMPLATE_HEADERS = ["Barcode", "Product Name", "Amazon", "Noon", "Noon Minutes", "Talabat Mart"]


def _normalise(header: str) -> str | None:
    key = str(header or "").strip().lower()
    for field, aliases in COLUMN_ALIASES.items():
        if key in aliases:
            return field
    return None


def _parse_ref(value: str) -> dict | None:
    """A cell is either a full product URL or a bare ASIN/SKU."""
    text = str(value or "").strip()
    if not text:
        return None
    if text.lower().startswith("http"):
        return {"url": text}
    # Platform modules look for the key they care about, so supply both.
    return {"asin": text, "sku": text}


def _rows_from_xlsx(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    mapping = {i: _normalise(h) for i, h in enumerate(header)}
    out = []
    for row in rows:
        out.append({mapping[i]: v for i, v in enumerate(row) if i in mapping and mapping[i]})
    wb.close()
    return out


def _rows_from_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        mapping = {h: _normalise(h) for h in (reader.fieldnames or [])}
        return [{mapping[h]: v for h, v in row.items() if mapping.get(h)} for row in reader]


def load_products(path: Path) -> dict:
    """Returns {barcode: {"name": str, "platforms": {platform_id: ref}}}."""
    if not path.exists():
        raise FileNotFoundError(
            f"{path} not found. Run `python -m scraper.local_run --init` to create a template."
        )

    rows = _rows_from_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") else _rows_from_csv(path)

    products = {}
    for row in rows:
        barcode = str(row.get("barcode") or "").strip()
        if not barcode or barcode.lower() == "none":
            continue
        # Excel often hands back numeric barcodes as floats (5056141881928.0).
        if barcode.endswith(".0"):
            barcode = barcode[:-2]

        amazon_ref = _parse_ref(row.get("amazon"))
        platforms = {}
        if amazon_ref:
            # One page load serves both Amazon tiers.
            platforms["amazon_core"] = amazon_ref
            platforms["amazon_fast"] = amazon_ref
        for field, platform_id in (
            ("noon", "noon_retail"),
            ("noon_minutes", "noon_minutes"),
            ("talabat", "talabat_mart"),
        ):
            ref = _parse_ref(row.get(field))
            if ref:
                platforms[platform_id] = ref

        if not platforms:
            continue

        products[barcode] = {"name": str(row.get("name") or "").strip(), "platforms": platforms}

    return products


def write_template(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append([
        "5056141881928",
        "Hotpack Bio-Degradable Garbage Bags 65x95cm (3x20pcs)",
        "B092BTLHZX",
        "https://www.noon.com/uae-en/hotpack-strong-bio-degradable-heavy-duty-disposable-garbage-20-pcs-65cm-x-95cm-pack-3-opgbr6595x3pkt-black-65x95cm/N45268893A/p/",
        "https://minutes.noon.com/uae-en/now-product/Z3A381CE007B28CE0D7AFZ-1/",
        "",
    ])

    widths = [18, 48, 22, 60, 60, 40]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
