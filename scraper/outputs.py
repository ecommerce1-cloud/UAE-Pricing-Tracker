"""Write results out as an Excel workbook and a standalone HTML dashboard.

The HTML file is fully self-contained (data, styles and logos all inlined) so it
opens by double-clicking with no server and no internet: a plain file:// page
cannot fetch sibling JSON, which is why nothing is loaded externally.
"""

import base64
import json
from datetime import datetime
from pathlib import Path

from .platforms import PLATFORMS
from .zones import ZONES

ROOT = Path(__file__).resolve().parent.parent
LOGO_DIR = ROOT / "docs" / "assets" / "logos"
TEMPLATE = Path(__file__).resolve().parent / "report_template.html"

LONG_HEADERS = [
    "Barcode", "Product Name", "Platform", "Zone", "Price", "Currency",
    "Available", "Source", "Note", "Checked At",
]


def _esc(value) -> str:
    text = str("" if value is None else value)
    text = text.replace("&", "&amp;").replace('"', "&quot;")
    return text.replace("<", "&lt;").replace(">", "&gt;")


def _zone_pairs(platform_result: dict):
    zones = platform_result.get("zones") or {}
    return [(z, zones.get(z["id"])) for z in ZONES]


def _live_prices(pairs) -> list:
    return [r["price"] for _, r in pairs if r and r.get("available") and r.get("price") is not None]


def _summary_cell(platform_result, zone_based: bool) -> str:
    if not platform_result:
        return ""
    if zone_based:
        prices = _live_prices(_zone_pairs(platform_result))
        if not prices:
            return ""
        lo, hi = min(prices), max(prices)
        return format(lo, ".2f") if lo == hi else format(lo, ".2f") + "-" + format(hi, ".2f")
    if platform_result.get("available") and platform_result.get("price") is not None:
        return format(platform_result["price"], ".2f")
    return ""


def long_rows(products: dict, results: dict, checked_at: str) -> list:
    """One row per barcode x platform x zone -- the full exportable detail."""
    rows = []
    for barcode, product in products.items():
        per_platform = results.get(barcode, {})
        for platform in PLATFORMS:
            res = per_platform.get(platform["id"])
            if not res:
                continue
            if platform["zone_based"]:
                for zone, r in _zone_pairs(res):
                    rows.append([
                        barcode, product.get("name", ""), platform["name"], zone["name"],
                        (r or {}).get("price"), (r or {}).get("currency", "AED"),
                        bool(r and r.get("available")), (r or {}).get("source") or "",
                        (r or {}).get("error") or "", checked_at,
                    ])
            else:
                rows.append([
                    barcode, product.get("name", ""), platform["name"], "",
                    res.get("price"), res.get("currency", "AED"),
                    bool(res.get("available")), res.get("source") or "",
                    res.get("error") or "", checked_at,
                ])
    return rows


def write_excel(path: Path, products: dict, results: dict, checked_at: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F232C")

    # Sheet 1: at-a-glance grid, barcode on the left and platforms across.
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Barcode", "Product Name"] + [p["name"] for p in PLATFORMS])
    ws.append(["", ""] + ["AED (min-max, 5 zones)" if p["zone_based"] else "AED" for p in PLATFORMS])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[2]:
        cell.font = Font(italic=True, size=8, color="808080")
        cell.alignment = Alignment(horizontal="center")

    for barcode, product in products.items():
        per_platform = results.get(barcode, {})
        summary = [_summary_cell(per_platform.get(p["id"]), p["zone_based"]) for p in PLATFORMS]
        ws.append([barcode, product.get("name", "")] + summary)

    for i, width in enumerate([18, 46] + [18] * len(PLATFORMS), start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "C3"

    # Sheet 2: every zone and every note, for analysis.
    detail = wb.create_sheet("By Zone")
    detail.append(LONG_HEADERS)
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
    for row in long_rows(products, results, checked_at):
        detail.append(row)
    for i, width in enumerate([18, 40, 16, 30, 10, 10, 10, 12, 60, 26], start=1):
        detail.column_dimensions[detail.cell(row=1, column=i).column_letter].width = width
    detail.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _logo_data_uri(filename: str) -> str:
    raw = (LOGO_DIR / filename).read_bytes()
    return "data:image/svg+xml;base64," + base64.b64encode(raw).decode()


def _na_span(error) -> str:
    text = str(error or "no data")
    blocked = text.startswith("blocked:")
    cls = "na blocked" if blocked else "na"
    glyph = "&#9888;" if blocked else "&mdash;"
    return '<span class="' + cls + '" title="' + _esc(text) + '">' + glyph + "</span>"


def _cell_html(result, zone_based: bool) -> str:
    if not result:
        return _na_span("not tracked on this platform")

    if not zone_based:
        if result.get("available") and result.get("price") is not None:
            return '<div class="price">AED ' + format(result["price"], ".2f") + "</div>"
        return _na_span(result.get("error"))

    pairs = _zone_pairs(result)
    prices = _live_prices(pairs)
    if not prices:
        first_error = next((r.get("error") for _, r in pairs if r and r.get("error")), "no data")
        return _na_span(first_error)

    lo, hi = min(prices), max(prices)
    if lo == hi:
        label = "AED " + format(lo, ".2f")
    else:
        label = "AED " + format(lo, ".2f") + "&ndash;" + format(hi, ".2f")

    breakdown = []
    for zone, r in pairs:
        if r and r.get("available") and r.get("price") is not None:
            value = "AED " + format(r["price"], ".2f")
        else:
            value = _na_span((r or {}).get("error"))
        breakdown.append("<div><span>" + _esc(zone["name"]) + "</span><span>" + value + "</span></div>")

    if len(prices) == len(ZONES):
        count = "5 zones"
    else:
        count = str(len(prices)) + "/" + str(len(ZONES)) + " zones"

    return (
        '<div class="price">' + label + "</div>"
        + '<span class="ztog">' + count + " &#9662;</span>"
        + '<div class="zbreak hide">' + "".join(breakdown) + "</div>"
    )


def write_html(path: Path, products: dict, results: dict, checked_at: str) -> None:
    head_cells = ""
    for p in PLATFORMS:
        head_cells += (
            '<th><div class="ph"><img src="' + _logo_data_uri(p["logo"])
            + '" alt="' + _esc(p["name"]) + '"/><span>' + _esc(p["name"])
            + "</span></div></th>"
        )

    body_rows = ""
    for barcode, product in products.items():
        per_platform = results.get(barcode, {})
        cells = ""
        for p in PLATFORMS:
            cells += '<td class="pc">' + _cell_html(per_platform.get(p["id"]), p["zone_based"]) + "</td>"
        name = _esc(product.get("name") or "(unnamed)")
        body_rows += (
            '<tr><td class="bc">' + _esc(barcode) + "</td>"
            + '<td><span class="pn">' + name + "</span></td>"
            + cells + "</tr>"
        )

    export_rows = [LONG_HEADERS]
    for row in long_rows(products, results, checked_at):
        export_rows.append(["" if v is None else v for v in row])

    stamp_dt = datetime.fromisoformat(checked_at)
    html = TEMPLATE.read_text(encoding="utf-8")
    tokens = (
        ("__STAMP__", stamp_dt.strftime("%d %b %Y, %H:%M")),
        ("__COUNT__", str(len(products))),
        ("__DATE__", stamp_dt.strftime("%Y-%m-%d")),
        ("__HEAD_CELLS__", head_cells),
        ("__BODY_ROWS__", body_rows),
        ("__ROWS_JSON__", json.dumps(export_rows)),
    )
    for token, value in tokens:
        html = html.replace(token, value)

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
